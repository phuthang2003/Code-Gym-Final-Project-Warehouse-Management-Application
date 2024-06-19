import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook, load_workbook
import os

class DrinkManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Drink Management")
        self.root.geometry("500x500")
        self.drinks = []

        # Load data from Excel file
        self.load_data()

        # Create UI components
        self.create_widgets()

        # Save data when the application exits
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def create_widgets(self):
        # Main Frame with background color
        main_frame = tk.Frame(self.root, bg="#f0f0f0", padx=10, pady=10)
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Labels and Entry fields
        tk.Label(main_frame, text="Drink Name", bg="#f0f0f0").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.drink_name_entry = tk.Entry(main_frame)
        self.drink_name_entry.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(main_frame, text="Price", bg="#f0f0f0").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.price_entry = tk.Entry(main_frame)
        self.price_entry.grid(row=1, column=1, padx=5, pady=5)

        tk.Label(main_frame, text="COGS", bg="#f0f0f0").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        self.cogs_entry = tk.Entry(main_frame)
        self.cogs_entry.grid(row=2, column=1, padx=5, pady=5)

        tk.Label(main_frame, text="Stock", bg="#f0f0f0").grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
        self.stock_entry = tk.Entry(main_frame)
        self.stock_entry.grid(row=3, column=1, padx=5, pady=5)

        # Buttons with custom styles and fixed size
        add_button = tk.Button(main_frame, text="Add Drink", bg="#4781f1", fg="white", font=("Montserrat", 8, "bold"), command=self.add_drink)
        add_button.grid(row=4, column=0, padx=5, pady=5, sticky=tk.EW)

        view_button = tk.Button(main_frame, text="View Drink", bg="#35b985", fg="white", font=("Montserrat", 8, "bold"), command=self.view_drink)
        view_button.grid(row=4, column=1, padx=5, pady=5, sticky=tk.EW)

        delete_button = tk.Button(main_frame, text="Delete Drink", bg="#e84547", fg="white", font=("Montserrat", 8, "bold"), command=self.delete_drink)
        delete_button.grid(row=5, column=0, padx=5, pady=5, sticky=tk.EW)

        reset_button = tk.Button(main_frame, text="Reset Fields", bg="#f09f2f", fg="white", font=("Montserrat", 8, "bold"), command=self.reset_fields)
        reset_button.grid(row=5, column=1, padx=5, pady=5, sticky=tk.EW)

        # Drink Treeview display
        self.tree = ttk.Treeview(main_frame, columns=("Name", "Price", "COGS", "Stock"), show="headings")
        self.tree.heading("Name", text="Name")
        self.tree.heading("Price", text="Price")
        self.tree.heading("COGS", text="COGS")
        self.tree.heading("Stock", text="Stock")
        self.tree.grid(row=6, column=0, columnspan=2, padx=5, pady=5, sticky=(tk.W, tk.E))

        self.scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.scrollbar.grid(row=6, column=2, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=self.scrollbar.set)

        # Update the Treeview with loaded data
        self.update_drink_list()

    def add_drink(self):
        name = self.drink_name_entry.get()
        price = self.price_entry.get()
        cogs = self.cogs_entry.get()
        stock = self.stock_entry.get()

        if name and price and cogs and stock:
            drink = {"name": name, "price": price, "cogs": cogs, "stock": stock}
            self.drinks.append(drink)
            print(f"Added drink: {drink}")  # Debugging statement
            self.update_drink_list()
            self.reset_fields()
            self.save_data()
        else:
            messagebox.showwarning("Input Error", "Please fill all fields")

    def view_drink(self):
        selected_item = self.tree.selection()
        if selected_item:
            item = self.tree.item(selected_item)
            drink = item['values']
            messagebox.showinfo("Drink Details", f"Name: {drink[0]}\nPrice: {drink[1]}\nCOGS: {drink[2]}\nStock: {drink[3]}")
        else:
            messagebox.showwarning("Selection Error", "Please select a drink to view")

    def delete_drink(self):
        selected_item = self.tree.selection()
        if selected_item:
            item = self.tree.item(selected_item)
            drink_name = item['values'][0]
            self.drinks = [drink for drink in self.drinks if drink["name"] != drink_name]
            print(f"Deleted drink: {drink_name}")  # Debugging statement
            self.update_drink_list()
            self.save_data()
        else:
            messagebox.showwarning("Selection Error", "Please select a drink to delete")

    def reset_fields(self):
        self.drink_name_entry.delete(0, tk.END)
        self.price_entry.delete(0, tk.END)
        self.cogs_entry.delete(0, tk.END)
        self.stock_entry.delete(0, tk.END)

    def update_drink_list(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for drink in self.drinks:
            self.tree.insert("", "end", values=(drink["name"], drink["price"], drink["cogs"], drink["stock"]))
        print(f"Updated Treeview with {len(self.drinks)} drinks")  # Debugging statement

    # Save data to a local Excel file
    def save_data(self):
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Name", "Price", "COGS", "Stock"])

            for drink in self.drinks:
                sheet.append([drink["name"], drink["price"], drink["cogs"], drink["stock"]])
                print(f"Saving drink: {drink}")  # Debugging statement

            workbook.save(r"D:\Python_Study\c0424l1-py103-project\Final_Project\drinks.xlsx")
            print("Data saved successfully")  # Debugging statement
        except Exception as e:
            print(f"Error saving data: {e}")

    # Load data from a local Excel file
    def load_data(self):
        if os.path.exists(r"D:\Python_Study\c0424l1-py103-project\Final_Project\drinks.xlsx"):
            try:
                workbook = load_workbook(r"D:\Python_Study\c0424l1-py103-project\Final_Project\drinks.xlsx")
                sheet = workbook.active

                self.drinks = []
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None:
                        drink = {"name": row[0], "price": row[1], "cogs": row[2], "stock": row[3]}
                        self.drinks.append(drink)
                        print(f"Loaded drink: {drink}")  # Debugging statement
                print("Data loaded successfully")  # Debugging statement
            except Exception as e:
                print(f"Error loading data: {e}")

    # Save data to a local Excel file before closing the application
    def on_closing(self):
        self.save_data()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = DrinkManagerApp(root)
    root.mainloop()
