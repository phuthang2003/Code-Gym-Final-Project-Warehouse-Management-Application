# Import các thư viện cần thiết
import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
import os

# Tạo class
class DrinkManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("VManage - Solution to warehouses")
        self.root.geometry("450x600")
        self.drinks = []

        # Load data từ file excel
        self.load_data()

        # Tạo các thành phần UI
        self.create_widgets()

        # Lưu dữ liệu khi đóng chương trình
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    # Function tạo các thành phần UI
    def create_widgets(self):
        # Setup khung để chứa các thành phần khác
        main_frame = tk.Frame(self.root, bg="#f0f0f0", padx=10, pady=10)
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Setup các Labels và khung điền thông tin

        # Labels và khung điền tên đồ uống
        tk.Label(main_frame, text="Drink Name", font=("Nunito", 10, "bold"), bg="#f0f0f0").grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        # Khung
        self.drink_name_entry = ttk.Entry(main_frame, width=20)
        # Vị trí khung
        self.drink_name_entry.grid(row=0, column=1, padx=0, pady=0)


        # Labels và khung điền giá
        tk.Label(main_frame, text="Price", font=("Nunito", 10, "bold"), bg="#f0f0f0").grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        # Khung
        self.price_entry = ttk.Entry(main_frame, width=20)
        # Vị trí khung
        self.price_entry.grid(row=1, column=1, padx=0, pady=0)


        # Labels và khung điền COGS
        tk.Label(main_frame, text="COGS", font=("Nunito", 10, "bold"), bg="#f0f0f0").grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        # Khung
        self.cogs_entry = ttk.Entry(main_frame, width=20)
        # Vị trí khung
        self.cogs_entry.grid(row=2, column=1, padx=0, pady=0)

        # Labels và khung điền stock
        tk.Label(main_frame, text="Stock", font=("Nunito", 10, "bold"), bg="#f0f0f0").grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
        # Khung
        self.stock_entry = ttk.Entry(main_frame, width=20)
        # Vị trí khung
        self.stock_entry.grid(row=3, column=1, padx=0, pady=0)

        # Khung điền số lượng bán
        # Khung
        self.sold_entry = ttk.Entry(main_frame, width=20)
        # Vị trí khung
        self.sold_entry.grid(row=5, column=1, padx=0, pady=0)

        # Quản lý hàng
        drink_info = tk.Label(main_frame, text="Inventory Management", font=("Montserrat", 8, "bold"), bg="lightgrey")
        drink_info.grid(row=4, column=0, columnspan=1, padx=5, pady=5)

        # Quản lý bán hàng
        sales_info = tk.Label(main_frame, text="Sales Management", font=("Montserrat", 8, "bold"), bg="lightgrey")
        sales_info.grid(row=4, column=1, columnspan=1, padx=5, pady=5)

        # Thêm các nút và format với các màu khác nhau
        # Nút add với màu xanh dương
        add_button = tk.Button(main_frame, text="Add Drink", bg="#4781f1", fg="white", font=("Nunito", 8, "bold"), command=self.add_drink)
        add_button.grid(row=5, column=0, padx=30, pady=5, sticky=tk.EW)

        # Nút view với màu xanh lá
        view_button = tk.Button(main_frame, text="View Drink", bg="#35b985", fg="white", font=("Nunito", 8, "bold"), command=self.view_drink)
        view_button.grid(row=6, column=0, padx=30, pady=5, sticky=tk.EW)

        # Nút deleve với màu đỏ
        delete_button = tk.Button(main_frame, text="Delete Drink", bg="#e84547", fg="white", font=("Nunito", 8, "bold"), command=self.delete_drink)
        delete_button.grid(row=7, column=0, padx=30, pady=5, sticky=tk.EW)

        # Nút reset với màu vàng
        reset_button = tk.Button(main_frame, text="Reset Fields", bg="#f09f2f", fg="white", font=("Nunito", 8, "bold"), command=self.reset_fields)
        reset_button.grid(row=8, column=0, padx=30, pady=5, sticky=tk.EW)

        # Nút bán màu cam
        sell_button = tk.Button(main_frame, text="Sell", bg="#f09f2f", fg="white", font=("Nunito", 8, "bold"), command=self.sell_drink)
        sell_button.grid(row=6, column=1, padx=30, pady=5, sticky=tk.EW)


        # Treeview để display các đồ uống được thêm vào
        self.tree = ttk.Treeview(main_frame, columns=("Name", "Price"), show="headings")
        # Cột tên
        self.tree.heading("Name", text="Name")
        # Cột giá
        self.tree.heading("Price", text="Price")
        # Setup vị trí cho bảng
        self.tree.grid(row=9, column=0, columnspan=2, padx=5, pady=5, sticky=(tk.W, tk.E))

        # Setup treeview để chứa thông tin về các đồ uống đã được nhập vào
        self.scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.scrollbar.grid(row=9, column=2, sticky=(tk.N, tk.S))
        self.tree.configure(yscrollcommand=self.scrollbar.set)
        
        # Upload data vào TreeView
        self.update_drink_list()

  
    # Setup các function cho chương trình
    # Add đồ uống
    def add_drink(self):
        name = self.drink_name_entry.get()
        price = int(self.price_entry.get())
        cogs = int(self.cogs_entry.get())
        stock = int(self.stock_entry.get())

        if name and price and cogs and stock: # Đảm bảo các thông tin đều được fill in đầy đủ
            drink = {"name": name, "price": price, "cogs": cogs, "stock": stock, "profit": 0}
            self.drinks.append(drink)
            self.update_drink_list()
            self.reset_fields() # Xóa các thông tin ở trong phần điền
        else:
            messagebox.showwarning("Input Error", "Please fill all fields")
    
    # View đồ uống
    def view_drink(self):
        selected_item = self.tree.selection()
        if selected_item: # Đảm bảo item được chọn không lỗi
            item = self.tree.item(selected_item)
            drink = item["values"]
            messagebox.showinfo("Drink Details", f"Name: {drink[0]}\nPrice: VND {drink[1]}\nCOGS: {drink[2]}\nStock: {drink[3]}\nProfit: VND {drink[4]}")
        else:
            messagebox.showwarning("Selection Error", "Please select a drink to view")

    # Xóa đồ uống
    def delete_drink(self):
        selected_item = self.tree.selection()
        if selected_item:
            item = self.tree.item(selected_item)
            drink_name = item["values"][0]
            self.drinks = [drink for drink in self.drinks if drink["name"] != drink_name]
            self.update_drink_list() # Update lại treeview với drinks mới
            self.save_data()
        else:
            messagebox.showwarning("Selection Error", "Please select a drink to delete")

    # Reset lại các field đã điền
    def reset_fields(self):
        self.drink_name_entry.delete(0, tk.END)
        self.price_entry.delete(0, tk.END)
        self.cogs_entry.delete(0, tk.END)
        self.stock_entry.delete(0, tk.END)

    # Update lại các field đã điền
    def update_drink_list(self):
        for item in self.tree.get_children(): # Trả về một list tất cả đồ uống trong tree
            self.tree.delete(item) # Xóa tất cả các đồ uống để treeview trống
        for drink in self.drinks:
            self.tree.insert("", "end", values=(drink["name"], drink["price"], drink["cogs"], drink["stock"], drink["profit"])) # Insert item từ vị trí đầu tiên, và item tiếp theo sẽ được insert ngay sau item trước đó
    
    # Function bán hàng
    def sell_drink(self):
        selected_item = self.tree.selection()
        if selected_item:
            item = self.tree.item(selected_item)
            drink_name = item["values"][0]
            number_sold = self.sold_entry.get()
            
            if number_sold.isdigit() and int(number_sold) > 0:
                number_sold = int(number_sold)
                for drink in self.drinks:
                    if drink["name"] == drink_name:
                        try:
                            # Ensure the stock is an integer
                            stock = int(drink["stock"])
                            price = int(drink["price"])
                            cogs = int(drink["cogs"])
                            
                            if stock >= number_sold:
                                drink["stock"] = stock - number_sold
                                profit = (price - cogs) * number_sold
                                drink["profit"] += profit
                                self.update_drink_list()
                                self.save_data()
                                messagebox.showinfo("Sell Drink", f"Sold {number_sold} of {drink_name}\nProfit: VND{profit}")
                                self.reset_fields()
                                self.number_sold_entry.delete(0, tk.END)
                                return
                            else:
                                messagebox.showwarning("Stock Error", "Not enough stock available")
                                return
                        except ValueError:
                            messagebox.showerror("Data Error", "Stock, Price, or COGS value is not an integer")
                            return
            else:
                messagebox.showwarning("Input Error", "Please enter a valid number of drinks sold")
        else:
            messagebox.showwarning("Selection Error", "Please select a drink to sell")
    

    # Lưu dữ liệu vào một local file Excel
    def save_data(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Name", "Price", "COGS", "Stock"]) # Thêm header vào sheet

        for drink in self.drinks: # Append dữ liệu từng drink vào sheet
            sheet.append([drink["name"], drink["price"], drink["cogs"], drink["stock"], drink["profit"]])

        workbook.save(r"D:\Python_Study\c0424l1-py103-project\Final_Project\drinks.xlsx")
    # Load dữ liệu từ một local file Excel
    def load_data(self):
        if os.path.exists(r"D:\Python_Study\c0424l1-py103-project\Final_Project\drinks.xlsx"):
            workbook = load_workbook(r"D:\Python_Study\c0424l1-py103-project\Final_Project\drinks.xlsx")
            sheet = workbook.active

            self.drinks = []
            for row in sheet.iter_rows(min_row=2, values_only=True): 
            # Lấy dữ liệu từ dòng thứ 2 vì dòng đầu là headers, và chỉ lấy dữ liệu nếu hàng đó chứa giá trị
            # Hàm for loop lấy dữ liêu từng dòng trong sheet
                if row[0] is not None: # Check cell đầu tiên trong row xem có giá trị hay không. Đảm bảo việc skip những row mà tên của đồ uống bị thiếu hoặc lỗi.
                    drink = {"name": row[0], "price": row[1], "cogs": row[2], "stock": row[3], "profit": row[4]} # Lấy lần lượt dữ liệu theo cell trong row
                    self.drinks.append(drink) # Append vào drinks chứa đồ uống
    
    # Lưu dữ liệu vào local file Excel trước khi đóng cửa sổ chương trình
    def on_closing(self):
        self.save_data()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = DrinkManagerApp(root)
    root.mainloop()
